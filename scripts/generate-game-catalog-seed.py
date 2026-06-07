from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_PLATFORM = "NS1"
DEFAULT_TIMESTAMP = "2026-06-07T00:00:00.000Z"

NAME_COLUMNS = ("游戏名", "游戏名称")
DATE_COLUMNS = ("发售日期",)
COVER_COLUMNS = ("封面原图URL",)
LOCAL_IMAGE_COLUMNS = ("入库用本地图片",)


def read_cell(row: dict[str, Any], columns: tuple[str, ...]) -> Any:
    for column in columns:
        if column in row:
            return row[column]
    return None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value))


def normalize_local_path(value: Any) -> str:
    return clean_text(value).replace("\\", "/")


def parse_release_date(value: Any) -> tuple[str | None, bool]:
    if value is None:
        return None, False

    if isinstance(value, datetime):
        return value.date().isoformat(), False

    raw = clean_text(value)
    if not raw:
        return None, False

    normalized = raw.replace(".", "/").replace("-", "/")
    for pattern in ("%Y/%m/%d", "%Y/%m", "%Y"):
        try:
            parsed = datetime.strptime(normalized, pattern)
            if pattern == "%Y/%m":
                return f"{parsed.year:04d}-{parsed.month:02d}-01", False
            if pattern == "%Y":
                return f"{parsed.year:04d}-01-01", False
            return parsed.date().isoformat(), False
        except ValueError:
            continue

    return raw, True


def create_catalog_id(platform: str, name: str) -> str:
    base = f"{platform}:{name}".casefold()
    digest = hashlib.sha1(base.encode("utf-8")).hexdigest()[:12]
    return f"catalog-{platform.lower()}-{digest}"


def load_rows(input_path: Path) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        return []

    headers = [clean_text(value) for value in raw_rows[0]]
    rows: list[tuple[int, dict[str, Any]]] = []
    for index, values in enumerate(raw_rows[1:], start=2):
        row = {header: value for header, value in zip(headers, values) if header}
        rows.append((index, row))
    return rows


def build_seed(input_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    items: list[dict[str, Any]] = []
    seen: set[str] = set()
    skipped_empty_name = 0
    duplicate_count = 0
    invalid_date_count = 0

    for source_row, row in load_rows(input_path):
        name = normalize_name(read_cell(row, NAME_COLUMNS))
        if not name:
            skipped_empty_name += 1
            continue

        platform = DEFAULT_PLATFORM
        dedupe_key = f"{platform}:{name}".casefold()
        if dedupe_key in seen:
            duplicate_count += 1
            continue
        seen.add(dedupe_key)

        release_date, invalid_date = parse_release_date(read_cell(row, DATE_COLUMNS))
        if invalid_date:
            invalid_date_count += 1

        cover_url = clean_text(read_cell(row, COVER_COLUMNS))
        local_image_path = normalize_local_path(read_cell(row, LOCAL_IMAGE_COLUMNS))

        item = {
            "id": create_catalog_id(platform, name),
            "name": name,
            "platform": platform,
            "source": "excel_import",
            "sourceRow": source_row,
            "createdAt": DEFAULT_TIMESTAMP,
            "updatedAt": DEFAULT_TIMESTAMP,
        }

        if release_date:
            item["releaseDate"] = release_date
        if cover_url:
            item["coverUrl"] = cover_url
            item["sourceOriginalUrl"] = cover_url
        if local_image_path:
            item["localImagePath"] = local_image_path

        items.append(item)

    summary = {
        "input": str(input_path),
        "totalRows": len(load_rows(input_path)),
        "imported": len(items),
        "skippedEmptyName": skipped_empty_name,
        "duplicates": duplicate_count,
        "invalidDates": invalid_date_count,
    }
    return items, summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate local game catalog seed JSON from games.xlsx.")
    parser.add_argument("--input", default="files/games.xlsx")
    parser.add_argument("--output", default="apps/miniprogram/data/game-catalog.seed.json")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    items, summary = build_seed(input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(items, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
